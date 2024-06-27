from openpyxl import Workbook
from openpyxl.styles import PatternFill, colors
import openpyxl
import csv

def main():
    file_name = 'maggio.xlsx'
    employees = ['Raffaele', 'Grazia', 'Nunzia', 'Roberta', 'Francesca', 'Viviana', 'Pouya', 'Chiara', 'Giacomo', 'Bianca']
    workbook = openpyxl.load_workbook(file_name)

    for ws in workbook:
        for x in range(1,500):
            for y in range(1,500):
                cell = ws.cell(row=x, column=y)
                if cell.value is None:
                    continue
                for idx, e in enumerate(employees):
                    if e in str(cell.value):
                        cell.fill = PatternFill(start_color=COLOR_INDEX[idx], end_color=COLOR_INDEX[idx], fill_type='solid')

    workbook.save(file_name)


COLOR_INDEX = (
    '00FFCC00', '003366FF', '00FF0000', '0000FF00', '0099CC00', #0-4
    '00FFFF00', '00FF00FF', '0000FFFF', '00FFCC99', '00CC99FF', #5-9
    '00FF0000', '0000FF00', '000000FF', '00FFFF00', '00FF00FF', #10-14
    '0000FFFF', '00800000', '00008000', '00000080', '00808000', #15-19
    '00800080', '00008080', '00C0C0C0', '00808080', '009999FF', #20-24
    '00993366', '00FFFFCC', '00CCFFFF', '00660066', '00FF8080', #25-29
    '000066CC', '00CCCCFF', '00000080', '00FF00FF', '00FFFF00', #30-34
    '0000FFFF', '00800080', '00800000', '00008080', '000000FF', #35-39
    '0000CCFF', '00CCFFFF', '00CCFFCC', '00FFFF99', '0099CCFF', #40-44
    '00FF99CC', '00CC99FF', '00FFCC99', '003366FF', '0033CCCC', #45-49
    '0099CC00', '00FFCC00', '00FF9900', '00FF6600', '00666699', #50-54
    '00969696', '00003366', '00339966', '00003300', '00333300', #55-59
    '00993300', '00993366', '00333399', '00333333',  #60-63
)
if __name__ == '__main__':
    main()

