from openpyxl import Workbook
from openpyxl.styles import PatternFill, colors
import openpyxl
import csv
import json
from pulp import LpStatus

MIP_TOLERANCE = 1e-6
FORMAT_DATE = "%d/%m/%Y"

def save_csv(solver, name_csv):
    from_date, to_date = solver.from_date, solver.to_date
    days, shift_types = solver.days, solver.shift_types
    employees, n_shifts = solver.employees, solver.n_shifts
    shifts, leave = solver.shifts, solver.leave
    map_slot_hours_t_i = solver.map_slot_hours_t_i
    #split_shift_emp = solver.split_shift_emp
    leave_gap_2_days = solver.leave_gap_2_days

    #print(solver.min_max_hours_emp)

    map_e_h = {}
    map_e_leave = {}
    map_e_leave_sat_sun = {}
    map_e_spezzati = {}
    map_e_split = {}
    data = []
    # init struct
    for i in range(90):
        data.append({})
    for e in employees:
        map_e_h[e] = {}
        map_e_leave[e] = {}
        map_e_leave_sat_sun[e] = 0
        map_e_split[e] = 0
        map_e_spezzati[e] = {}
        map_e_h[e]['total'] = 0
        map_e_leave[e]['total'] = 0

        weeks = solver.get_weeks_between_dates(from_date, to_date)
        for w in weeks:
            map_e_h[e][w] = 0
            map_e_leave[e][w] = 0
        for d in days:
            map_e_spezzati[e][d] = 0
    data[0]['Summary'] = ''
    
    # loop
    for d in days:
        j = 0
        # add slots
        for i in solver.get_indexes_shift(d):
            for e in employees:
                if abs(1 - shifts[d][i][e].varValue) <= MIP_TOLERANCE:
                    data[j][f'{d}:Giorno'] = d
                    data[j][f'{d}:Da'] = f'{map_slot_hours_t_i[d][i].t_from.strftime("%H:%M")} - {map_slot_hours_t_i[d][i].t_to.strftime("%H:%M")}'
                    data[j][f'{d}:Durata H'] = map_slot_hours_t_i[d][i].duration
                    data[j][f'{d}:Tipo'] = map_slot_hours_t_i[d][i].type
                    
                    data[j][f'{d}:Nome'] = e
                    
                    week = map_slot_hours_t_i[d][i].week
                    map_e_h[e]['total'] = map_slot_hours_t_i[d][i].duration + map_e_h[e]['total']
                    map_e_h[e][week] = map_slot_hours_t_i[d][i].duration + map_e_h[e][week]
                    j+=1

        # add leave days
        for e in employees:
            if abs(1 - leave[d][e].varValue) <= MIP_TOLERANCE:
                data[j][f'{d}:Giorno'] = d
                data[j][f'{d}:Da'] = 'Riposo'
                data[j][f'{d}:Nome'] = e


                week = map_slot_hours_t_i[d][i].week
                map_e_leave[e]['total'] = map_e_leave[e]['total'] + 1
                map_e_leave[e][week] = map_e_leave[e][week] + 1
                if map_slot_hours_t_i[d][0].day_of_week in [6,7]:
                    map_e_leave_sat_sun[e] = map_e_leave_sat_sun[e] + 1
                j+=1

    for e in employees:
        lastday = -1
        for d in days:
            for i in solver.get_indexes_shift(d):
                if abs(1 - shifts[d][i][e].varValue) <= MIP_TOLERANCE:
                    if lastday == d:
                        map_e_split[e] = map_e_split[e] + 1
                    lastday = d

    # Add total, week hours for employee
    j+=3
    data[j]['Summary'] = 'Ore totali'
    j+=1
    for k,v in map_e_h.items():
        data[j]['Summary'] = k + ' ' + str(v)
        j+=1

    # Add total leave days for employee
    j+=2
    data[j]['Summary'] = 'Ferie totali'
    j+=1
    for k,v in map_e_leave.items():
        #gap_2_days = sum(leave_gap_2_days[d][k].varValue if leave_gap_2_days[d][k].varValue is not None else 0 for d in days[:-1] ) 
        data[j]['Summary'] = '{}: {} ({}) '.format(k, v, map_e_leave_sat_sun[k])
        j+=1
    
    j+=2
    data[j]['Summary'] = 'Spezzati totali'
    j+=1
    for k,v in map_e_split.items():
        data[j]['Summary'] = k + ' ' + str(v)
        j+=1

    fieldnames = data[0].keys()
    with open(name_csv, 'w') as myfile:
        wr = csv.DictWriter(myfile,  fieldnames=fieldnames)
        wr.writeheader()
        wr.writerows(data)

def response_build(solver):
    from_date, to_date = solver.from_date, solver.to_date
    days, shift_types = solver.days, solver.shift_types
    employees, n_shifts = solver.employees, solver.n_shifts
    shifts, leave = solver.shifts, solver.leave
    map_slot_hours_t_i = solver.map_slot_hours_t_i
    #split_shift_emp = solver.split_shift_emp
    leave_gap_2_days = solver.leave_gap_2_days


    response = {} 
    scheduling = {}
    response['status'] = LpStatus[solver.status] # Not Solved, Optimal, Infeasible, Unbounded, Undefined
    response['from_date'] = solver.from_date.strftime(FORMAT_DATE)
    response['to_date'] = solver.to_date.strftime(FORMAT_DATE)
    response['num_days'] = solver.num_days
    response['employees'] = solver.employees
    response['employees_far'] = solver.employees_far
    response['scheduling'] = scheduling
    # loop
    for d in days:
        j = 0
        # add slots
        for i in solver.get_indexes_shift(d):
            for e in employees:
                if abs(1 - shifts[d][i][e].varValue) <= MIP_TOLERANCE:
                    date_str = map_slot_hours_t_i[d][i].date.strftime(FORMAT_DATE)
                    shift = {}
                    shift['day'] = map_slot_hours_t_i[d][i].date.strftime(FORMAT_DATE)
                    shift['from'] = f'{map_slot_hours_t_i[d][i].t_from.strftime("%H:%M")}'
                    shift['to'] = f'{map_slot_hours_t_i[d][i].t_to.strftime("%H:%M")}'
                    shift['duration'] = map_slot_hours_t_i[d][i].duration
                    shift['type'] = map_slot_hours_t_i[d][i].type
                    shift['id'] = map_slot_hours_t_i[d][i].id
                    shift['period'] = map_slot_hours_t_i[d][i].period
                    shift['employee'] = e
                    if date_str not in scheduling:
                         scheduling[date_str] = []
                    scheduling[date_str].append(shift)              

        # add leave days
        for e in employees:
            if abs(1 - leave[d][e].varValue) <= MIP_TOLERANCE:
                    date_str = map_slot_hours_t_i[d][i].date.strftime(FORMAT_DATE)
                    shift = {}
                    shift['day'] = map_slot_hours_t_i[d][0].date.strftime(FORMAT_DATE)
                    shift['from'] = ''
                    shift['to'] = ''
                    shift['duration'] = ''
                    shift['type'] = 'Riposo'
                    shift['id'] = -1
                    shift['period'] = 'LV'
                    shift['employee'] = e
                    if date_str not in scheduling:
                         scheduling[date_str] = []
                    scheduling[date_str].append(shift)   
    
    return response

def save_excel(solver, name_csv, name_excel):
    COLOR_INDEX = (
        '00FFCC00', '003366FF', '00FF0000', '0000FF00', '0099CC00', #0-4
        '00FFFF00', '00FF00FF', '0000FFFF', '00FFCC99', '00CC99FF', #5-9
        '00FF0000', '0000FF00', '000000FF', '00FFFF00', '00FF00FF', #10-14
        '0000FFFF', '00800000', '00008000', '00000080', '00808000', #15-19
        '00800080', '00008080', '00C0C0C0', '00808080', '009999FF', #20-24
        '00993366', '00FFFFCC', '00CCFFFF', '00660066', '00FF8080', #25-29
        '000066CC', '00CCCCFF', '00000080', '00FF00FF', '00FFFF00', #30-34
        '0000FFFF', '00800080', '00800000', '00008080', '000000FF', #35-39
        '0000CCFF', '00CCFFFF', '00CCFFCC', '00FFFF99', '0099CCFF', #40-44
        '00FF99CC', '00CC99FF', '00FFCC99', '003366FF', '0033CCCC', #45-49
        '0099CC00', '00FFCC00', '00FF9900', '00FF6600', '00666699', #50-54
        '00969696', '00003366', '00339966', '00003300', '00333300', #55-59
        '00993300', '00993366', '00333399', '00333333',  #60-63
    )

    wb = openpyxl.Workbook()
    worksheet = wb.active

    with open(name_csv) as f:
        reader = csv.reader(f, delimiter=',')
        for row in reader:
            worksheet.append(row)

    for x in range(1,500):
        for y in range(1,500):
            cell = worksheet.cell(row=x, column=y)
            if cell.value is None:
                continue
            for idx, e in enumerate(solver.employees):
                if e in str(cell.value):
                    cell.fill = PatternFill(start_color=COLOR_INDEX[idx], end_color=COLOR_INDEX[idx], fill_type='solid')

    wb.save(name_excel)
